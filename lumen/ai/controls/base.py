from __future__ import annotations

import asyncio
import io
import json
import pathlib
import zipfile

import pandas as pd
import param

from panel.layout import HSpacer
from panel.pane.markup import HTML
from panel.viewable import Viewer
from panel_material_ui import (
    Button, Column, Column as MuiColumn, IconButton, Markdown,
    RadioButtonGroup, Row, Select, Tabs, TextInput,
)

from ...sources.duckdb import DuckDBSource
from ...util import detect_file_encoding, normalize_table_name
from ..utils import log_debug
from .progress import Progress

TABLE_EXTENSIONS = ("csv", "parquet", "parq", "json", "xlsx", "geojson", "wkt", "zip")

METADATA_EXTENSIONS = ("md", "txt", "yaml", "yml", "json", "pdf", "docx", "doc", "pptx", "ppt")
METADATA_FILENAME_PATTERNS = ("_metadata", "metadata_", "readme", "schema")

# Download configuration constants
class DownloadConfig:
    CHUNK_SIZE = 1024 * 1024  # 1MB chunks
    TIMEOUT_SECONDS = 300     # 5 minutes
    PROGRESS_UPDATE_INTERVAL = 50  # Update every 50 chunks
    DEFAULT_HASH_MODULO = 10000
    UNKNOWN_SIZE_MAX = 1000000000  # 1GB max for unknown file sizes

    # Connection settings
    CONNECTION_LIMIT = 100
    CONNECTION_LIMIT_PER_HOST = 30
    KEEPALIVE_TIMEOUT = 30


# ─────────────────────────────────────────────────────────────────────────────
# RESULT TYPES
# ─────────────────────────────────────────────────────────────────────────────

class SourceResult(param.Parameterized):
    """
    Result of a source loading operation.

    This is the return type for `_load()` methods in source controls.
    Use the factory methods for common patterns.
    """

    sources = param.List(default=[], doc="List of data sources loaded.")

    table = param.String(default=None, allow_None=True, doc="Primary table name.")

    metadata = param.Dict(default={}, doc="Additional metadata about the loaded data.")

    message = param.String(default=None, allow_None=True, doc="Status message to display.")

    @classmethod
    def from_dataframe(
        cls,
        df: pd.DataFrame,
        table_name: str,
        message: str | None = None,
        **metadata
    ) -> SourceResult:
        """
        Create result from a DataFrame.
        """
        source = DuckDBSource.from_df(tables={table_name: df})
        source.tables[table_name] = f"SELECT * FROM {table_name}"
        if metadata:
            source.metadata = {table_name: metadata}
        return cls(
            sources=[source],
            table=table_name,
            metadata=metadata,
            message=message or f"Loaded {len(df):,} rows into '{table_name}'"
        )

    @classmethod
    def from_source(
        cls,
        source: DuckDBSource,
        table: str | None = None,
        message: str | None = None,
    ) -> SourceResult:
        """
        Create result from an existing source.
        """
        return cls(sources=[source], table=table, message=message)

    @classmethod
    def empty(cls, message: str = "No data loaded") -> SourceResult:
        """
        Return an empty result.
        """
        return cls(message=message)


# ─────────────────────────────────────────────────────────────────────────────
# FILE ROW COMPONENT
# ─────────────────────────────────────────────────────────────────────────────

class UploadedFileRow(Viewer):
    """
    Row for a single uploaded file with data/metadata classification.

    Displays each file as a row where users can toggle it as data or metadata.
    """

    filename = param.String(default="", doc="Filename without extension")

    extension = param.String(default="", doc="File extension")

    file_type = param.Selector(default="data", objects=["data", "metadata"], doc="""
        Classification of the file. 'data' files are processed as tables,
        'metadata' files are added to the document vector store.""")

    alias = param.String(default="", doc="Table alias (for data files only)")

    sheet = param.Selector(default=None, objects=[], doc="Sheet (for xlsx files)")

    delete = param.Event(doc="Delete this card")

    _load_sheets = param.Event(doc="Load sheets")

    def __init__(self, file_obj: io.BytesIO, **params):
        filename = params.get("filename", "")
        extension = params.get("extension", "")

        # Parse filename if contains extension
        if "/" in filename:
            filename = filename.rsplit("/")[-1]
        if "." in filename and not extension:
            filename, extension = filename.rsplit(".", maxsplit=1)

        params["filename"] = filename
        params["extension"] = extension.lower()
        params["alias"] = normalize_table_name(filename)

        # Auto-detect file type
        is_metadata = (
            extension.lower() in METADATA_EXTENSIONS or
            any(pattern in filename.lower() for pattern in METADATA_FILENAME_PATTERNS)
        )
        params.setdefault("file_type", "metadata" if is_metadata else "data")

        super().__init__(**params)
        self.file_obj = file_obj

        # Build UI components
        self._build_ui()

    def _build_ui(self):
        self._type_toggle = RadioButtonGroup.from_param(
            self.param.file_type,
            options={"data": "data", "metadata": "metadata"},
            size="small",
            label="",
            margin=(0, 10),
            align="center",
        )

        self._delete_button = IconButton.from_param(
            self.param.delete,
            icon="close",
            size="small",
            margin=(0, 5)
        )

        self._alias_input = TextInput.from_param(
            self.param.alias,
            placeholder="Table alias",
            size="small",
            visible=self.file_type == "data",
            margin=10,
            width=200,
            align="center"
        )

        self._sheet_select = Select.from_param(
            self.param.sheet,
            name="Sheet",
            size="small",
            visible=False,
            margin=(5, 10)
        )

        # Trigger sheet loading for xlsx
        if self.extension == "xlsx":
            self.param.trigger("_load_sheets")

    @param.depends("file_type", watch=True)
    def _update_alias_visibility(self):
        if hasattr(self, '_alias_input'):
            self._alias_input.visible = self.file_type == "data"

    @param.depends("_load_sheets", watch=True)
    async def _handle_load_sheets(self):
        """Load sheet names for xlsx files."""
        if self.extension != "xlsx":
            return
        import openpyxl
        wb = openpyxl.load_workbook(self.file_obj, read_only=True)
        self.param.sheet.objects = wb.sheetnames
        self.sheet = wb.sheetnames[0]
        self._sheet_select.visible = True
        self.file_obj.seek(0)  # Reset file pointer

    @param.depends("alias", watch=True)
    def _sanitize_alias(self):
        """Ensure alias is a valid SQL identifier.

        Uses the same normalization as DuckDBSource.normalize_table.
        """
        sanitized = normalize_table_name(self.alias)
        if sanitized != self.alias:
            self.alias = sanitized

    def __panel__(self):
        filename_display = Markdown(
            f"`{self.filename}.{self.extension}`",
            margin=(5, 10),
            styles={"flex-shrink": "0"}
        )

        main_row = Row(
            filename_display,
            self._type_toggle,
            self._alias_input,
            HSpacer(),
            self._delete_button,
            sizing_mode="stretch_width",
            margin=(5, 0),
        )

        if self.extension == "xlsx":
            return Column(
                main_row,
                self._sheet_select,
                margin=0,
                sizing_mode="stretch_width"
            )
        else:
            return main_row


# ─────────────────────────────────────────────────────────────────────────────
# BASE SOURCE CONTROLS
# ─────────────────────────────────────────────────────────────────────────────

class BaseSourceControls(Viewer):
    """
    Base class for source controls with lifecycle management.
    """

    context = param.Dict(default={})

    disabled = param.Boolean(default=False, doc="Disable controls")

    load_mode = param.Selector(
        default="button",
        objects=["button", "manual"],
        constant=True,
        doc="""
        How data loading is triggered:
        - "button": Show load button, clicking triggers _load()
        - "manual": No button, use _run_load(coro) for custom triggers
        """
    )

    multiple = param.Boolean(default=True, doc="Allow multiple files")

    clear_uploads = param.Boolean(default=True, doc="Clear uploaded file tabs")

    replace_controls = param.Boolean(default=False, doc="Replace controls on add")

    filedropper_kwargs = param.Dict(default={}, doc="""Keyword arguments to pass to FileDropper.
        Common options include 'accepted_filetypes' and 'max_file_size'.
        See https://panel.holoviz.org/reference/widgets/FileDropper.html for all available options.""")

    outputs = param.Dict(default={})

    source_catalog = param.Parameter(default=None, doc="Reference to SourceCatalog instance")

    upload_handlers = param.Dict(default={}, doc="Handlers for custom file extensions")

    # Events
    add = param.Event(doc="Use uploaded file(s)")

    cancel = param.Event(doc="Cancel")

    load = param.Event(doc="Trigger data loading")

    upload_successful = param.Event(doc="Triggered when files are successfully uploaded and processed")

    # Internal state
    _last_table = param.String(default="", doc="Last table added")

    _count = param.Integer(default=0, doc="Count of sources added")

    # UI customization
    label = param.String(default="", constant=True, doc="""
        HTML label shown in the sidebar for this control.""" )

    load_button_label = param.String(default="Load Data", doc="""
        Text label for the load button.""")

    load_button_icon = param.String(default="download", doc="""
        Material icon name for the load button.""")

    source_name_prefix = param.String(default="UploadedSource", doc="""
        Prefix for auto-generated source names.""")

    __abstract = True

    def __init__(self, **params):
        super().__init__(**params)
        self._markitdown = None
        self._file_cards = []
        self._init_ui_components()
        self._layout = self._render_layout()

    def _init_ui_components(self):
        """Initialize shared UI components. Called before _render_layout()."""
        self._error_placeholder = HTML("", visible=False, margin=(0, 10, 5, 10))
        self._message_placeholder = HTML("", visible=False, margin=(0, 10, 10, 10))

        self._error_text = self._error_placeholder
        self._message_text = self._message_placeholder

        self.progress = self._render_progress()

        self._upload_cards = MuiColumn(
            sizing_mode="stretch_width",
            margin=0,
            styles={"border-top": "1px solid #e0e0e0", "padding-top": "5px"}
        )
        self._upload_cards.visible = False

        files_to_process = self._upload_cards.param["objects"].rx.len() > 0
        self._add_button = Button.from_param(
            self.param.add,
            name="Confirm file(s)",
            icon="add",
            visible=files_to_process,
            description="",
            align="center",
            sizing_mode="stretch_width",
            height=42,
        )

        self._load_button = Button.from_param(
            self.param.load,
            label=self.param.load_button_label,
            icon=self.param.load_button_icon,
            sizing_mode="stretch_width",
            description="",
            height=42,
        )

        self.tables_tabs = Tabs(sizing_mode="stretch_width")

    async def _load(self) -> SourceResult:
        raise NotImplementedError(
            f"{self.__class__.__name__} must implement _load() or set "
            "load_mode='manual' and use _run_load() for custom triggers."
        )

    def _render_controls(self) -> list:
        return []

    def _render_progress(self) -> Progress:
        return Progress()

    def _render_layout(self) -> Viewer:
        controls = self._render_controls()
        components = list(controls)
        if self.load_mode == "button":
            components.append(self._load_button)

        components.extend([
            self._error_placeholder,
            self._message_placeholder,
            self.progress.bar,
            self.progress.description,
        ])

        return MuiColumn(
            *components,
            sizing_mode="stretch_width",
            margin=(10, 15),
        )

    @param.depends("load", watch=True)
    async def _on_load_event(self):
        await self._run_load(self._load())

    async def _run_load(self, coro):
        self._clear_messages()
        self._layout.loading = True
        self._load_button.disabled = True

        try:
            await asyncio.sleep(0.01)
            result = await coro
            self._handle_success(result)
        except Exception as e:
            self._handle_error(e)
            raise
        finally:
            self._layout.loading = False
            self._load_button.disabled = False
            self.progress.clear()

    def _clear_messages(self):
        self._error_placeholder.visible = False
        self._message_placeholder.visible = False
        self._error_placeholder.object = ""
        self._message_placeholder.object = ""

    def _handle_success(self, result: SourceResult):
        if not result.sources:
            self._show_message(result.message or "No data loaded", error=True)
            return

        for source in result.sources:
            self.outputs["source"] = source
            if "sources" not in self.outputs:
                self.outputs["sources"] = []
            self.outputs["sources"].append(source)

        if result.table:
            self.outputs["table"] = result.table
            self._last_table = result.table

        self.param.trigger("outputs")
        self.param.trigger("upload_successful")
        self._show_message(result.message or "✓ Data loaded successfully")

    def _handle_error(self, error: Exception):
        self._show_message(f"⚠️ {error}", error=True)

    def _show_message(self, message: str, error: bool = False):
        target = self._error_placeholder if error else self._message_placeholder
        target.object = message
        target.visible = True

    def _create_file_object(self, file_data: bytes | io.BytesIO | io.StringIO, suffix: str):
        if isinstance(file_data, (io.BytesIO, io.StringIO)):
            return file_data

        if suffix == "csv" and isinstance(file_data, bytes):
            encoding = detect_file_encoding(file_data)
            file_data = file_data.decode(encoding).encode("utf-8")
        return io.BytesIO(file_data) if isinstance(file_data, bytes) else io.StringIO(file_data)

    def _format_bytes(self, bytes_size):
        if bytes_size == 0:
            return "0 B"
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        while bytes_size >= 1024 and i < len(size_names) - 1:
            bytes_size /= 1024.0
            i += 1
        return f"{bytes_size:.1f} {size_names[i]}"

    def _generate_file_cards(self, files: dict):
        self._upload_cards.clear()
        self._file_cards.clear()

        if not files:
            self._add_button.visible = False
            self._upload_cards.visible = False
            return

        for filename, file_data in files.items():
            suffix = pathlib.Path(filename).suffix.lstrip(".").lower()
            file_obj = self._create_file_object(file_data, suffix)

            card = UploadedFileRow(
                file_obj=file_obj,
                filename=filename
            )
            card.param.watch(lambda e, c=card: self._remove_card(c), "delete")

            self._upload_cards.append(card)
            self._file_cards.append(card)

        self._upload_cards.visible = len(self._file_cards) > 0
        self._add_button.visible = len(self._file_cards) > 0

    def _remove_card(self, card: UploadedFileRow):
        if card in self._file_cards:
            self._file_cards.remove(card)
        if card in self._upload_cards.objects:
            self._upload_cards.remove(card)

        self._add_button.visible = len(self._file_cards) > 0
        self._upload_cards.visible = len(self._file_cards) > 0

    def _add_table(
        self,
        duckdb_source: DuckDBSource,
        file: io.BytesIO | io.StringIO,
        card: UploadedFileRow,
    ) -> int:
        conn = duckdb_source._connection
        extension = card.extension
        table = card.alias
        sql_expr = f"SELECT * FROM {table}"
        params = {}
        conversion = None
        filename = f"{card.filename}.{extension}"

        try:
            file.seek(0)
            if extension.endswith("csv"):
                df = pd.read_csv(file, parse_dates=True, sep=None, engine='python')
            elif extension.endswith(("parq", "parquet")):
                df = pd.read_parquet(file)
            elif extension.endswith("json"):
                df = self._read_json_file(file, filename)
            elif extension.endswith("xlsx"):
                sheet = card.sheet
                df = pd.read_excel(file, sheet_name=sheet)
            elif extension.endswith(('geojson', 'wkt', 'zip')):
                df, conversion, params = self._read_geo_file(file, extension, table, conn)
            else:
                self._error_placeholder.object += f"\\n⚠️ Could not convert {filename!r}: unsupported format."
                self._error_placeholder.visible = True
                return 0
        except Exception as e:
            self._error_placeholder.object += f"\\n⚠️ Error processing {filename!r}: {e}"
            self._error_placeholder.visible = True
            return 0

        if df is None or df.empty:
            self._error_placeholder.object += f"\\n⚠️ {filename!r} contains no data."
            self._error_placeholder.visible = True
            return 0

        # Convert pandas StringDtype columns to object for DuckDB compatibility
        for col in df.columns:
            if isinstance(df[col].dtype, pd.StringDtype):
                df[col] = df[col].astype(object)

        duckdb_source.param.update(params)
        df_rel = conn.from_df(df)
        if conversion:
            conn.register(f'{table}_temp', df_rel)
            conn.execute(conversion)
            conn.unregister(f'{table}_temp')
        else:
            df_rel.to_view(table)
        duckdb_source.tables[table] = sql_expr
        self.outputs["source"] = duckdb_source
        if "sources" not in self.outputs:
            self.outputs["sources"] = [duckdb_source]
        else:
            self.outputs["sources"].append(duckdb_source)
        self.outputs["table"] = table
        self.param.trigger('outputs')
        self._last_table = table
        return 1

    def _read_json_file(self, file: io.BytesIO | io.StringIO, filename: str) -> pd.DataFrame:
        file.seek(0)
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8')

        content = content.strip()
        if not content:
            raise ValueError("JSON file is empty")

        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            raise ValueError("Invalid JSON") from e

        if isinstance(data, list):
            if len(data) == 0:
                raise ValueError("JSON array is empty")
            if all(isinstance(item, dict) for item in data):
                return pd.DataFrame(data)
            else:
                return pd.DataFrame({"value": data})
        elif isinstance(data, dict):
            data_keys = ['data', 'records', 'rows', 'items', 'results']
            for key in data_keys:
                if key in data and isinstance(data[key], list):
                    if len(data[key]) > 0 and all(isinstance(item, dict) for item in data[key]):
                        return pd.DataFrame(data[key])
            if all(isinstance(v, list) for v in data.values()):
                lengths = [len(v) for v in data.values()]
                if len(set(lengths)) == 1:
                    return pd.DataFrame(data)
                else:
                    raise ValueError(f"JSON object has arrays of different lengths: {lengths}")
            if all(not isinstance(v, (list, dict)) or v is None for v in data.values()):
                return pd.DataFrame([data])
            file.seek(0)
            try:
                return pd.read_json(file)
            except ValueError as e:
                raise ValueError("JSON structure is not tabular.") from e
        else:
            raise ValueError(f"JSON root must be an object or array, got {type(data).__name__}")

    def _read_geo_file(self, file: io.BytesIO, extension: str, table: str, conn) -> tuple[pd.DataFrame, str | None, dict]:
        if extension.endswith('zip'):
            zf = zipfile.ZipFile(file)
            if not any(f.filename.endswith('shp') for f in zf.filelist):
                raise ValueError("ZIP file does not contain a shapefile (.shp)")
            file.seek(0)

        import geopandas as gpd
        geo_df = gpd.read_file(file)

        if geo_df.empty:
            raise ValueError("Geospatial file contains no features")

        df = pd.DataFrame(geo_df)
        df['geometry'] = geo_df['geometry'].to_wkb()

        params = {
            'initializers': [
                """
                INSTALL spatial;
                LOAD spatial;
                """
            ]
        }
        conn.execute(params['initializers'][0])

        cols = ', '.join(f'"{c}"' for c in df.columns if c != 'geometry')
        conversion = f'CREATE TEMP TABLE {table} AS SELECT {cols}, ST_GeomFromWKB(geometry) as geometry FROM {table}_temp'

        return df, conversion, params

    def _extract_metadata_content(self, file_obj: io.BytesIO, extension: str) -> str:
        file_obj.seek(0)
        if extension in ("md", "txt", "yaml", "yml"):
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            return content
        elif extension == "json":
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            try:
                data = json.loads(content)
                return json.dumps(data, indent=2)
            except json.JSONDecodeError:
                return content
        else:
            from markitdown import MarkItDown
            if self._markitdown is None:
                self._markitdown = MarkItDown()
            return self._markitdown.convert_stream(file_obj, file_extension=extension).text_content

    def _add_metadata_file(self, card: UploadedFileRow) -> int:
        try:
            content = self._extract_metadata_content(card.file_obj, card.extension)
            base_filename = f"{card.filename}.{card.extension}"
            filename = base_filename
            if self.source_catalog:
                existing = [m["filename"] for m in self.source_catalog._available_metadata]
                if filename in existing:
                    name_without_ext = card.filename
                    ext = card.extension
                    counter = 1
                    while filename in existing:
                        filename = f"{name_without_ext}_{counter}.{ext}"
                        counter += 1
                    self._message_placeholder.param.update(
                        object=f"Renamed duplicate: {base_filename} → {filename}",
                        visible=True
                    )
                metadata_entry = {
                    "filename": filename,
                    "display_name": filename.rsplit('.', 1)[0],
                    "content": content
                }
                # Preserve raw bytes for PDF files so they can be rendered natively
                if card.extension == "pdf":
                    card.file_obj.seek(0)
                    metadata_entry["raw_bytes"] = card.file_obj.read()
                self.source_catalog._available_metadata.append(metadata_entry)
                asyncio.create_task(self._sync_metadata(filename))  # noqa: RUF006

            self.param.trigger('outputs')
            return 1
        except Exception as e:
            self._error_placeholder.object += f"\\nCould not process metadata file {card.filename}: {e}"
            self._error_placeholder.visible = True
            return 0

    async def _sync_metadata(self, filename: str):
        if not self.source_catalog:
            return
        await self.source_catalog._sync_metadata_to_vector_store(filename)

    def _process_files(self):
        self._error_placeholder.object = ""
        self._error_placeholder.visible = False

        if len(self._file_cards) == 0:
            return 0, 0, 0

        source = None
        n_tables = 0
        n_metadata = 0
        table_upload_callbacks = {
            key.lstrip("."): value
            for key, value in self.upload_handlers.items()
        }
        custom_table_extensions = tuple(table_upload_callbacks)

        data_cards = [c for c in self._file_cards if c.file_type == "data"]
        metadata_cards = [c for c in self._file_cards if c.file_type == "metadata"]

        for card in data_cards:
            log_debug(f"Processing data card: {card.filename}.{card.extension} (alias: {card.alias})")
            if card.extension.endswith(custom_table_extensions):
                source = table_upload_callbacks[card.extension](
                    self.context, card.file_obj, card.alias, card.filename
                )
                if source is not None:
                    n_tables += len(source.get_tables())
                    self.outputs["source"] = source
                    if "sources" not in self.outputs:
                        self.outputs["sources"] = [source]
                    else:
                        self.outputs["sources"].append(source)
                    self.param.trigger("outputs")
            elif card.extension.endswith(TABLE_EXTENSIONS):
                if source is None:
                    source_id = f"{self.source_name_prefix}{self._count:06d}"
                    source = DuckDBSource(uri=":memory:", ephemeral=True, name=source_id, tables={})
                table_name = card.alias
                filename = f"{card.filename}.{card.extension}"
                if table_name not in source.metadata:
                    source.metadata[table_name] = {}
                source.metadata[table_name]["filename"] = filename
                n_tables += self._add_table(source, card.file_obj, card)
            else:
                self._error_placeholder.object += f"\\n⚠️ Skipped '{card.filename}.{card.extension}': unsupported format."
                self._error_placeholder.visible = True

        for card in metadata_cards:
            n_metadata += self._add_metadata_file(card)

        log_debug(f"Processed files: {n_tables} tables, {n_metadata} metadata files")
        return n_tables, 0, n_metadata

    def _clear_uploads(self):
        self._upload_cards.clear()
        self._file_cards.clear()
        self._add_button.visible = False

    def __panel__(self):
        return self._layout
