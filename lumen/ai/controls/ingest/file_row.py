from __future__ import annotations

import io

import param

from panel.layout import HSpacer
from panel.viewable import Viewer
from panel_material_ui import (
    Column, IconButton, Markdown, RadioButtonGroup, Row, Select, TextInput,
)

from ....util import normalize_table_name
from .constants import METADATA_EXTENSIONS, METADATA_FILENAME_PATTERNS


class UploadedFileRow(Viewer):
    """
    Row for a single uploaded file with data/metadata classification.

    Displays each file as a row where users can toggle it as data or metadata,
    set a table alias, and (for xlsx files) choose a sheet.
    """

    filename = param.String(default="", doc="Filename without extension")

    extension = param.String(default="", doc="File extension")

    file_type = param.Selector(default="data", objects=["data", "metadata"], doc="""
        Classification of the file. 'data' files are processed as tables,
        'metadata' files are added to the document vector store.""")

    alias = param.String(default="", doc="Table alias (for data files only)")

    sheet = param.Selector(default=None, objects=[], doc="Sheet (for xlsx files)")

    delete = param.Event(doc="Delete this card")

    _load_sheets = param.Event(doc="Load sheets")

    def __init__(self, file_obj: io.BytesIO, **params):
        filename = params.get("filename", "")
        extension = params.get("extension", "")

        # Strip directory prefix if present
        if "/" in filename:
            filename = filename.rsplit("/")[-1]
        # Split extension from filename if not provided separately
        if "." in filename and not extension:
            filename, extension = filename.rsplit(".", maxsplit=1)

        params["filename"] = filename
        params["extension"] = extension.lower()
        params["alias"] = normalize_table_name(filename)

        # Auto-detect file type from extension and filename patterns
        is_metadata = (
            extension.lower() in METADATA_EXTENSIONS or
            any(pattern in filename.lower() for pattern in METADATA_FILENAME_PATTERNS)
        )
        params.setdefault("file_type", "metadata" if is_metadata else "data")

        super().__init__(**params)
        self.file_obj = file_obj
        self._build_ui()

    def _build_ui(self):
        self._type_toggle = RadioButtonGroup.from_param(
            self.param.file_type,
            options={"data": "data", "metadata": "metadata"},
            size="small",
            label="",
            margin=(0, 10),
            align="center",
        )

        self._delete_button = IconButton.from_param(
            self.param.delete,
            icon="close",
            size="small",
            margin=(0, 5),
        )

        self._alias_input = TextInput.from_param(
            self.param.alias,
            placeholder="Table alias",
            size="small",
            visible=self.file_type == "data",
            margin=10,
            width=200,
            align="center",
        )

        self._sheet_select = Select.from_param(
            self.param.sheet,
            name="Sheet",
            size="small",
            visible=False,
            margin=(5, 10),
        )

        if self.extension == "xlsx":
            self.param.trigger("_load_sheets")

    @param.depends("file_type", watch=True)
    def _update_alias_visibility(self):
        if hasattr(self, "_alias_input"):
            self._alias_input.visible = self.file_type == "data"

    @param.depends("_load_sheets", watch=True)
    async def _handle_load_sheets(self):
        """Load sheet names for xlsx files."""
        if self.extension != "xlsx":
            return
        import openpyxl
        wb = openpyxl.load_workbook(self.file_obj, read_only=True)
        self.param.sheet.objects = wb.sheetnames
        self.sheet = wb.sheetnames[0]
        self._sheet_select.visible = True
        self.file_obj.seek(0)

    @param.depends("alias", watch=True)
    def _sanitize_alias(self):
        """Ensure alias is a valid SQL identifier."""
        sanitized = normalize_table_name(self.alias)
        if sanitized != self.alias:
            self.alias = sanitized

    def __panel__(self):
        filename_display = Markdown(
            f"`{self.filename}.{self.extension}`",
            margin=(5, 10),
            styles={"flex-shrink": "0"},
        )

        main_row = Row(
            filename_display,
            self._type_toggle,
            self._alias_input,
            HSpacer(),
            self._delete_button,
            sizing_mode="stretch_width",
            margin=(5, 0),
        )

        if self.extension == "xlsx":
            return Column(
                main_row,
                self._sheet_select,
                margin=0,
                sizing_mode="stretch_width",
            )
        return main_row
