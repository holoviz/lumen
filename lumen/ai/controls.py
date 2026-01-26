from __future__ import annotations

import asyncio
import io
import pathlib
import re
import zipfile

from typing import TYPE_CHECKING
from urllib.parse import parse_qs, urlparse

import aiohttp
import pandas as pd
import param

from panel.io import state
from panel.layout import Column, HSpacer, Row
from panel.pane.markup import HTML, Markdown
from panel.viewable import Viewer
from panel.widgets import FileDropper
from panel_material_ui import (
    AutocompleteInput, Button, Card, Column as MuiColumn, IconButton,
    LinearProgress, Popup, RadioButtonGroup, Select, Tabs, TextAreaInput,
    TextInput, ToggleIcon, Tree, Typography,
)

from ..config import load_yaml
from ..pipeline import Pipeline
from ..sources.duckdb import DuckDBSource
from ..util import detect_file_encoding, normalize_table_name
from .config import SOURCE_TABLE_SEPARATOR
from .context import TContext
from .utils import generate_diff, log_debug

if TYPE_CHECKING:
    from .views import SQLOutput

TABLE_EXTENSIONS = ("csv", "parquet", "parq", "json", "xlsx", "geojson", "wkt", "zip")

METADATA_EXTENSIONS = ("md", "txt", "yaml", "yml", "json")
METADATA_FILENAME_PATTERNS = ("_metadata", "metadata_", "readme", "schema")

# Download configuration constants
class DownloadConfig:
    CHUNK_SIZE = 1024 * 1024  # 1MB chunks
    TIMEOUT_SECONDS = 300     # 5 minutes
    PROGRESS_UPDATE_INTERVAL = 50  # Update every 50 chunks
    DEFAULT_HASH_MODULO = 10000
    UNKNOWN_SIZE_MAX = 1000000000  # 1GB max for unknown file sizes

    # Connection settings
    CONNECTION_LIMIT = 100
    CONNECTION_LIMIT_PER_HOST = 30
    KEEPALIVE_TIMEOUT = 30


class UploadedFileRow(Viewer):
    """
    Row for a single uploaded file with data/metadata classification.

    Displays each file as a row where users can toggle it as data or metadata.
    """

    filename = param.String(default="", doc="Filename without extension")

    extension = param.String(default="", doc="File extension")

    file_type = param.Selector(default="data", objects=["data", "metadata"], doc="""
        Classification of the file. 'data' files are processed as tables,
        'metadata' files are added to the document vector store.""")

    alias = param.String(default="", doc="Table alias (for data files only)")

    sheet = param.Selector(default=None, objects=[], doc="Sheet (for xlsx files)")

    delete = param.Event(doc="Delete this card")

    _load_sheets = param.Event(doc="Load sheets")

    def __init__(self, file_obj: io.BytesIO, **params):
        filename = params.get("filename", "")
        extension = params.get("extension", "")

        # Parse filename if contains extension
        if "/" in filename:
            filename = filename.rsplit("/")[-1]
        if "." in filename and not extension:
            filename, extension = filename.rsplit(".", maxsplit=1)

        params["filename"] = filename
        params["extension"] = extension.lower()
        params["alias"] = normalize_table_name(filename)

        # Auto-detect file type
        is_metadata = (
            extension.lower() in METADATA_EXTENSIONS or
            any(pattern in filename.lower() for pattern in METADATA_FILENAME_PATTERNS)
        )
        params.setdefault("file_type", "metadata" if is_metadata else "data")

        super().__init__(**params)
        self.file_obj = file_obj

        # Build UI components
        self._build_ui()

    def _build_ui(self):
        self._type_toggle = RadioButtonGroup.from_param(
            self.param.file_type,
            options={"data": "data", "metadata": "metadata"},
            size="small",
            label="",
            margin=(0, 10),
            align="center",
        )

        self._delete_button = IconButton.from_param(
            self.param.delete,
            icon="close",
            size="small",
            margin=(0, 5)
        )

        self._alias_input = TextInput.from_param(
            self.param.alias,
            placeholder="Table alias",
            size="small",
            visible=self.file_type == "data",
            margin=10,
            width=200,
            align="center"
        )

        self._sheet_select = Select.from_param(
            self.param.sheet,
            name="Sheet",
            size="small",
            visible=False,
            margin=(5, 10)
        )

        # Trigger sheet loading for xlsx
        if self.extension == "xlsx":
            self.param.trigger("_load_sheets")

    @param.depends("file_type", watch=True)
    def _update_alias_visibility(self):
        if hasattr(self, '_alias_input'):
            self._alias_input.visible = self.file_type == "data"

    @param.depends("_load_sheets", watch=True)
    async def _handle_load_sheets(self):
        """Load sheet names for xlsx files."""
        if self.extension != "xlsx":
            return
        import openpyxl
        wb = openpyxl.load_workbook(self.file_obj, read_only=True)
        self.param.sheet.objects = wb.sheetnames
        self.sheet = wb.sheetnames[0]
        self._sheet_select.visible = True
        self.file_obj.seek(0)  # Reset file pointer

    @param.depends("alias", watch=True)
    def _sanitize_alias(self):
        """Ensure alias is a valid SQL identifier.

        Uses the same normalization as DuckDBSource.normalize_table.
        """
        sanitized = normalize_table_name(self.alias)
        if sanitized != self.alias:
            self.alias = sanitized

    def __panel__(self):
        filename_display = Markdown(
            f"`{self.filename}.{self.extension}`",
            margin=(5, 10),
            styles={"flex-shrink": "0"}
        )

        main_row = Row(
            filename_display,
            self._type_toggle,
            self._alias_input,
            HSpacer(),
            self._delete_button,
            sizing_mode="stretch_width",
            margin=(5, 0),
        )

        if self.extension == "xlsx":
            return Column(
                main_row,
                self._sheet_select,
                margin=0,
                sizing_mode="stretch_width"
            )
        else:
            return main_row


class BaseSourceControls(Viewer):
    """
    Base class for source controls providing shared functionality for
    processing uploaded/downloaded files into data sources.
    """

    context = param.Dict(default={})

    disabled = param.Boolean(default=False, doc="Disable controls")

    add = param.Event(doc="Use uploaded file(s)")

    cancel = param.Event(doc="Cancel")

    multiple = param.Boolean(default=True, doc="Allow multiple files")

    clear_uploads = param.Boolean(default=True, doc="Clear uploaded file tabs")

    replace_controls = param.Boolean(default=False, doc="Replace controls on add")

    upload_successful = param.Event(doc="Triggered when files are successfully uploaded and processed")

    outputs = param.Dict(default={})

    source_catalog = param.Parameter(default=None, doc="Reference to SourceCatalog instance")

    upload_handlers = param.Dict(default={}, doc="Handlers for custom file extensions")

    _last_table = param.String(default="", doc="Last table added")

    _count = param.Integer(default=0, doc="Count of sources added")

    label = ""

    __abstract = True

    def __init__(self, **params):
        super().__init__(**params)
        self._markitdown = None
        self._file_cards = []  # Track UploadedFileRow instances

        # Shared UI components
        self._upload_cards = MuiColumn(sizing_mode="stretch_width", margin=0, styles={"border-top": "1px solid #e0e0e0", "padding-top": "5px"})
        self._upload_cards.visible = False

        files_to_process = self._upload_cards.param["objects"].rx.len() > 0
        self._add_button = Button.from_param(
            self.param.add,
            name="Confirm file(s)",
            icon="add",
            visible=files_to_process,
            description="",
            align="center",
            sizing_mode="stretch_width",
            height=42,
        )

        self._error_placeholder = HTML("", visible=False, margin=(0, 10, 5, 10))
        self._message_placeholder = HTML("", visible=False, margin=(0, 10, 10, 10))

        # Progress bar for operations
        self._progress_description = Typography(
            styles={"margin-left": "auto", "margin-right": "auto"},
            visible=False
        )
        self._progress_bar = LinearProgress(
            visible=False,
            margin=(0, 10, 5, 10),
            sizing_mode="stretch_width"
        )

        self.tables_tabs = Tabs(sizing_mode="stretch_width")

    def _create_file_object(self, file_data: bytes | io.BytesIO | io.StringIO, suffix: str):
        """Create appropriate file object based on file type"""
        if isinstance(file_data, (io.BytesIO, io.StringIO)):
            return file_data

        if suffix == "csv" and isinstance(file_data, bytes):
            encoding = detect_file_encoding(file_data)
            file_data = file_data.decode(encoding).encode("utf-8")
        return io.BytesIO(file_data) if isinstance(file_data, bytes) else io.StringIO(file_data)

    def _format_bytes(self, bytes_size):
        """Convert bytes to human readable format"""
        if bytes_size == 0:
            return "0 B"
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        while bytes_size >= 1024 and i < len(size_names) - 1:
            bytes_size /= 1024.0
            i += 1
        return f"{bytes_size:.1f} {size_names[i]}"

    def _generate_file_cards(self, files: dict):
        """Generate file cards for uploaded/downloaded files"""
        self._upload_cards.clear()
        self._file_cards.clear()

        if not files:
            self._add_button.visible = False
            self._upload_cards.visible = False
            return

        for filename, file_data in files.items():
            suffix = pathlib.Path(filename).suffix.lstrip(".").lower()
            file_obj = self._create_file_object(file_data, suffix)

            card = UploadedFileRow(
                file_obj=file_obj,
                filename=filename
            )
            card.param.watch(lambda e, c=card: self._remove_card(c), "delete")

            self._upload_cards.append(card)
            self._file_cards.append(card)

        self._upload_cards.visible = len(self._file_cards) > 0
        self._add_button.visible = len(self._file_cards) > 0

    def _remove_card(self, card: UploadedFileRow):
        """Remove a file card from the upload list."""
        if card in self._file_cards:
            self._file_cards.remove(card)
        if card in self._upload_cards.objects:
            self._upload_cards.remove(card)

        self._add_button.visible = len(self._file_cards) > 0
        self._upload_cards.visible = len(self._file_cards) > 0

    def _add_table(
        self,
        duckdb_source: DuckDBSource,
        file: io.BytesIO | io.StringIO,
        card: UploadedFileRow,
    ) -> int:
        conn = duckdb_source._connection
        extension = card.extension
        table = card.alias
        sql_expr = f"SELECT * FROM {table}"
        params = {}
        conversion = None
        if extension.endswith("csv"):
            df = pd.read_csv(file, parse_dates=True, sep=None, engine='python')
        elif extension.endswith(("parq", "parquet")):
            df = pd.read_parquet(file)
        elif extension.endswith("json"):
            df = pd.read_json(file)
        elif extension.endswith("xlsx"):
            sheet = card.sheet
            df = pd.read_excel(file, sheet_name=sheet)
        elif extension.endswith(('geojson', 'wkt', 'zip')):
            if extension.endswith('zip'):
                zf = zipfile.ZipFile(file)
                if not any(f.filename.endswith('shp') for f in zf.filelist):
                    raise ValueError("Could not interpret zip file contents")
                file.seek(0)
            import geopandas as gpd
            geo_df = gpd.read_file(file)
            df = pd.DataFrame(geo_df)
            df['geometry'] = geo_df['geometry'].to_wkb()
            params['initializers'] = init = ["""
            INSTALL spatial;
            LOAD spatial;
            """]
            conn.execute(init[0])
            cols = ', '.join(f'"{c}"' for c in df.columns if c != 'geometry')
            conversion = f'CREATE TEMP TABLE {table} AS SELECT {cols}, ST_GeomFromWKB(geometry) as geometry FROM {table}_temp'
        else:
            self._error_placeholder.object += f"\nCould not convert {card.filename}.{extension}."
            self._error_placeholder.visible = True
            return 0

        duckdb_source.param.update(params)
        df_rel = conn.from_df(df)
        if conversion:
            conn.register(f'{table}_temp', df_rel)
            conn.execute(conversion)
            conn.unregister(f'{table}_temp')
        else:
            df_rel.to_view(table)
        duckdb_source.tables[table] = sql_expr
        self.outputs["source"] = duckdb_source
        if "sources" not in self.outputs:
            self.outputs["sources"] = [duckdb_source]
        else:
            self.outputs["sources"].append(duckdb_source)
        self.outputs["table"] = table
        self.param.trigger('outputs')
        self._last_table = table
        return 1

    def _extract_metadata_content(self, file_obj: io.BytesIO, extension: str) -> str:
        """Extract text content from a metadata file."""
        file_obj.seek(0)

        if extension in ("md", "txt", "yaml", "yml"):
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            return content
        elif extension == "json":
            import json
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')
            try:
                data = json.loads(content)
                return json.dumps(data, indent=2)
            except json.JSONDecodeError:
                return content
        else:
            if self._markitdown is None:
                from markitdown import MarkItDown
                self._markitdown = MarkItDown()
            return self._markitdown.convert_stream(file_obj, file_extension=extension).text_content

    def _add_metadata_file(self, card: UploadedFileRow) -> int:
        """
        Add a metadata file for later association with tables.
        Stores in the SourceCatalog instance if available.

        Parameters
        ----------
        card : UploadedFileRow
            The file card containing the metadata file.
        """
        try:
            content = self._extract_metadata_content(card.file_obj, card.extension)
            base_filename = f"{card.filename}.{card.extension}"

            # Handle duplicate filenames with counter suffix
            filename = base_filename
            if self.source_catalog:
                existing = [m["filename"] for m in self.source_catalog._available_metadata]
                if filename in existing:
                    # Find unique filename with counter suffix
                    name_without_ext = card.filename
                    ext = card.extension
                    counter = 1
                    while filename in existing:
                        filename = f"{name_without_ext}_{counter}.{ext}"
                        counter += 1

                    # Inform user about the rename
                    self._message_placeholder.param.update(
                        object=f"Renamed duplicate: {base_filename} → {filename}",
                        visible=True
                    )

                self.source_catalog._available_metadata.append({
                    "filename": filename,
                    "display_name": filename.rsplit('.', 1)[0],
                    "content": content
                })
                # Sync to vector store via SourceCatalog
                asyncio.create_task(self._sync_metadata(filename))  # noqa: RUF006

            self.param.trigger('outputs')
            return 1
        except Exception as e:
            self._error_placeholder.object += f"\nCould not process metadata file {card.filename}: {e}"
            self._error_placeholder.visible = True
            return 0

    async def _sync_metadata(self, filename: str):
        """
        Sync metadata to vector store.
        """
        if not self.source_catalog:
            return

        # Sync to vector store
        await self.source_catalog._sync_metadata_to_vector_store(filename)

    def _process_files(self):
        """Process all file cards and add them as sources."""
        # Clear previous error/warning messages
        self._error_placeholder.object = ""
        self._error_placeholder.visible = False

        if len(self._file_cards) == 0:
            return 0, 0, 0

        source = None
        n_tables = 0
        n_metadata = 0
        table_upload_callbacks = {
            key.lstrip("."): value
            for key, value in self.upload_handlers.items()
        }
        custom_table_extensions = tuple(table_upload_callbacks)

        # Separate data and metadata files
        data_cards = [c for c in self._file_cards if c.file_type == "data"]
        metadata_cards = [c for c in self._file_cards if c.file_type == "metadata"]

        # Process data files (tables)
        for card in data_cards:
            log_debug(f"Processing data card: {card.filename}.{card.extension} (alias: {card.alias})")
            if card.extension.endswith(custom_table_extensions):
                log_debug(f"Using custom handler for extension: {card.extension}")
                source = table_upload_callbacks[card.extension](
                    self.context, card.file_obj, card.alias, card.filename
                )
                if source is not None:
                    n_tables += len(source.get_tables())
                    self.outputs["source"] = source
                    if "sources" not in self.outputs:
                        self.outputs["sources"] = [source]
                    else:
                        self.outputs["sources"].append(source)
                    self.param.trigger("outputs")
            elif card.extension.endswith(TABLE_EXTENSIONS):
                log_debug(f"Processing as table with extension: {card.extension}")
                if source is None:
                    source_id = f"UploadedSource{self._count:06d}"
                    source = DuckDBSource(uri=":memory:", ephemeral=True, name=source_id, tables={})
                table_name = card.alias
                filename = f"{card.filename}.{card.extension}"
                if table_name not in source.metadata:
                    source.metadata[table_name] = {}
                source.metadata[table_name]["filename"] = filename
                n_tables += self._add_table(source, card.file_obj, card)
            else:
                log_debug(f"Skipping file with unrecognized extension: {card.filename}.{card.extension}. Valid extensions: {TABLE_EXTENSIONS}")
                self._error_placeholder.object += f"\n⚠️ Skipped '{card.filename}.{card.extension}': unsupported format. Supported: {', '.join(TABLE_EXTENSIONS)}."
                self._error_placeholder.visible = True

        # Process metadata files
        for card in metadata_cards:
            n_metadata += self._add_metadata_file(card)

        log_debug(f"Processed files: {n_tables} tables, {n_metadata} metadata files")
        return n_tables, 0, n_metadata

    def _clear_uploads(self):
        """Clear uploaded files from view."""
        self._upload_cards.clear()
        self._file_cards.clear()
        self._add_button.visible = False

    def __panel__(self):
        raise NotImplementedError("Subclasses must implement __panel__")


class UploadControls(BaseSourceControls):
    """
    Controls for uploading files from the local filesystem.
    """

    label = '<span class="material-icons" style="vertical-align: middle;">upload</span> Upload Data'

    def __init__(self, **params):
        super().__init__(**params)

        self._file_input = FileDropper(
            layout="compact",
            multiple=self.param.multiple,
            margin=1,
            sizing_mode="stretch_width",
            disabled=self.param.disabled,
            stylesheets=[".bk-input.filepond--root { box-shadow: unset; cursor: grab; } .bk-input.filepond--root:not([disabled]):hover { box-shadow: unset; }"],
            visible=self._upload_cards.param.visible.rx.not_()
        )
        self._file_input.param.watch(self._on_file_upload, "value")

        self._layout = MuiColumn(
            self._file_input,
            self._upload_cards,
            self._add_button,
            self._error_placeholder,
            self._message_placeholder,
            self._progress_bar,
            self._progress_description,
            sizing_mode="stretch_width",
        )

    def _on_file_upload(self, event):
        """Handle file upload from FileDropper."""
        self._generate_file_cards(self._file_input.value or {})

    @param.depends("add", watch=True)
    def _on_add(self):
        """Process uploaded files."""
        if len(self._file_cards) == 0:
            return

        with self._layout.param.update(loading=True):
            n_tables, n_docs, n_metadata = self._process_files()

            total_files = len(self._file_cards)
            if self.clear_uploads:
                self._clear_uploads()
                self._file_input.value = {}

            if (n_tables + n_docs + n_metadata) > 0:
                self._message_placeholder.param.update(
                    object=f"Successfully processed {total_files} files ({n_tables} table(s), {n_metadata} metadata file(s)).",
                    visible=True,
                )
            self._error_placeholder.object = self._error_placeholder.object.strip()

        self._count += 1

        if (n_tables + n_docs + n_metadata) > 0:
            self.param.trigger('upload_successful')

    def __panel__(self):
        return self._layout


class DownloadControls(BaseSourceControls):
    """
    Controls for downloading files from URLs.
    """

    download_url = param.String(default="", doc="Enter one or more URLs, one per line, and press <Shift+Enter> to download", label="Download URL(s)")

    input_placeholder = param.String(default="Enter URLs, one per line, and press <Shift+Enter> to download")

    _active_download_task = param.ClassSelector(class_=asyncio.Task)

    label = '<span class="material-icons" style="vertical-align: middle;">download</span> Fetch Remote Data'

    def __init__(self, **params):
        super().__init__(**params)

        self._url_input = TextAreaInput.from_param(
            self.param.download_url,
            placeholder=self.param.input_placeholder,
            rows=2,
            margin=10,
            sizing_mode="stretch_width",
            disabled=self.param.disabled,
        )
        self._url_input.param.watch(self._handle_urls, "enter_pressed")

        self._cancel_button = Button.from_param(
            self.param.cancel,
            name="Cancel",
            icon="close",
            on_click=self._handle_cancel,
            visible=self.param._active_download_task.rx.is_not(None),
            height=42,
        )

        self._active_download_task = None

        self._layout = MuiColumn(
            self._url_input,
            self._upload_cards,
            Row(self._add_button, self._cancel_button),
            self._error_placeholder,
            self._message_placeholder,
            self._progress_bar,
            self._progress_description,
            sizing_mode="stretch_width",
        )

    def _handle_cancel(self, event):
        """Handle cancel button click by cancelling active download task"""
        if self._active_download_task and not self._active_download_task.done():
            self._active_download_task.cancel()
            self._progress_bar.visible = False
            self._message_placeholder.param.update(
                object="Download cancelled by user.",
                visible=True
            )
        # Clear the active task so Cancel button hides
        self._active_download_task = None

    def _is_valid_url(self, url):
        """Basic URL validation"""
        try:
            result = urlparse(url.strip())
            return all([result.scheme, result.netloc])
        except Exception:
            return False

    def _extract_filename_from_url(self, url: str) -> str:
        """Extract filename from URL with fallback logic"""
        parsed = urlparse(url)
        filename = parsed.path.split('/')[-1] if parsed.path else 'downloaded_file'

        if '?' in filename:
            filename = filename.split('?')[0]

        # Check if the extracted extension is a valid table/metadata extension
        current_ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        valid_extensions = TABLE_EXTENSIONS + METADATA_EXTENSIONS

        # If extension is not valid, check for format= query parameter
        if current_ext not in valid_extensions:
            query_params = parse_qs(parsed.query)
            format_param = query_params.get('format', [None])[0]
            if format_param and format_param.lower() in valid_extensions:
                # Replace or add the correct extension
                base_name = filename.rsplit('.', 1)[0] if '.' in filename else filename
                filename = f"{base_name}.{format_param.lower()}"

        if not filename or '.' not in filename:
            filename = f"data_{abs(hash(url)) % DownloadConfig.DEFAULT_HASH_MODULO}.json"

        return filename

    def _extract_filename_from_headers(self, response_headers: dict, default_filename: str) -> str:
        """Extract filename from HTTP headers if available"""
        # First check Content-Disposition header
        if 'content-disposition' in response_headers:
            cd = response_headers['content-disposition']
            matches = re.findall('filename="?([^"]+)"?', cd)
            if matches:
                suggested_name = matches[0]
                if '.' in suggested_name:
                    return suggested_name

        # Check if current filename has invalid extension, try to fix from Content-Type
        current_ext = default_filename.rsplit('.', 1)[-1].lower() if '.' in default_filename else ''
        valid_extensions = TABLE_EXTENSIONS + METADATA_EXTENSIONS

        if current_ext not in valid_extensions and 'content-type' in response_headers:
            content_type = response_headers['content-type'].lower().split(';')[0].strip()
            content_type_map = {
                'text/csv': 'csv',
                'application/csv': 'csv',
                'application/json': 'json',
                'application/geo+json': 'geojson',
                'application/vnd.geo+json': 'geojson',
                'application/parquet': 'parquet',
                'application/vnd.apache.parquet': 'parquet',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': 'xlsx',
                'application/vnd.ms-excel': 'xlsx',
                'text/plain': 'csv',  # Often CSVs are served as text/plain
            }
            if content_type in content_type_map:
                ext = content_type_map[content_type]
                base_name = default_filename.rsplit('.', 1)[0] if '.' in default_filename else default_filename
                return f"{base_name}.{ext}"

        return default_filename

    def _setup_progress_bar(self, content_length: str) -> int:
        """Setup progress bar based on content length"""
        self._progress_description.visible = True
        if content_length:
            total_size = int(content_length)
            self._progress_bar.param.update(
                visible=True,
                value=0,
                variant="determinate"
            )
            return total_size
        else:
            self._progress_bar.variant = "indeterminate"
            self._progress_bar.visible = True
            return 0

    def _update_download_progress(self, filename: str, downloaded_size: int, total_size: int, chunk_count: int):
        """Update progress bar if it's time for an update"""
        if chunk_count % DownloadConfig.PROGRESS_UPDATE_INTERVAL == 0:
            if total_size > 0:
                progress = min((total_size / downloaded_size) * 100, 100)
                if progress < 100:
                    self._progress_bar.value = progress
                    progress_desc = f"Downloading {filename}: {self._format_bytes(downloaded_size)}/{self._format_bytes(total_size)}"
                else:
                    progress_desc = f"Downloading {filename}: {self._format_bytes(downloaded_size)}"
            else:
                progress_desc = f"Downloading {filename}: {self._format_bytes(downloaded_size)}"

            self._progress_description.object = progress_desc

    def _finalize_download_progress(self, filename: str, downloaded_size: int, total_size: int):
        """Show final progress update when download completes"""
        if total_size > 0:
            progress = min((downloaded_size / total_size) * 100, 100)
            self._progress_bar.value = progress
            self._progress_description.object = f"Downloaded {filename}: {self._format_bytes(downloaded_size)}/{self._format_bytes(total_size)}"
        else:
            self._progress_bar.value = 100
            self._progress_description.object = f"Downloaded {filename}: {self._format_bytes(downloaded_size)}"

    def _format_download_error(self, error: Exception) -> str:
        """Format download error messages consistently"""
        if isinstance(error, aiohttp.ClientError):
            return f"Network error: {error!s}"
        elif isinstance(error, asyncio.TimeoutError):
            return f"Download timeout ({DownloadConfig.TIMEOUT_SECONDS}s exceeded)"
        elif isinstance(error, asyncio.CancelledError):
            return "Download cancelled"
        else:
            return f"Download failed: {error!s}"

    async def _download_file(self, url):
        """Download file from URL with progress bar and return (filename, file_content, error)"""
        try:
            filename = self._extract_filename_from_url(url)

            timeout = aiohttp.ClientTimeout(total=DownloadConfig.TIMEOUT_SECONDS)
            connector = aiohttp.TCPConnector(
                limit=DownloadConfig.CONNECTION_LIMIT,
                limit_per_host=DownloadConfig.CONNECTION_LIMIT_PER_HOST,
                keepalive_timeout=DownloadConfig.KEEPALIVE_TIMEOUT,
                enable_cleanup_closed=True
            )

            async with aiohttp.ClientSession(timeout=timeout, connector=connector) as session:
                async with session.get(url) as response:
                    response.raise_for_status()

                    filename = self._extract_filename_from_headers(response.headers, filename)
                    content_length = response.headers.get('content-length')
                    total_size = self._setup_progress_bar(content_length)

                    downloaded_data = io.BytesIO()
                    downloaded_size = 0
                    chunk_count = 0
                    size_mismatch_detected = False

                    async for chunk in response.content.iter_chunked(DownloadConfig.CHUNK_SIZE):
                        if asyncio.current_task().cancelled():
                            raise asyncio.CancelledError()

                        downloaded_data.write(chunk)
                        downloaded_size += len(chunk)
                        chunk_count += 1

                        if not size_mismatch_detected and total_size > 0 and downloaded_size > total_size:
                            # Content-Length was inaccurate, switch to indeterminate mode
                            total_size = 0
                            size_mismatch_detected = True

                        self._update_download_progress(filename, downloaded_size, total_size, chunk_count)
                        await asyncio.sleep(0)

                    self._finalize_download_progress(filename, downloaded_size, total_size)

                    downloaded_data.seek(0)
                    content = downloaded_data.read()

                    self._progress_bar.visible = False
                    self._progress_description.visible = False
                    return filename, content, None

        except (TimeoutError, aiohttp.ClientError, asyncio.CancelledError, Exception) as e:
            self._progress_bar.visible = False
            self._progress_description.visible = False
            return None, None, self._format_download_error(e)

    def _handle_urls(self, event):
        """Handle URL input and trigger downloads"""
        url_text = self._url_input.value
        if not url_text:
            return

        urls = [line.strip() for line in url_text.split('\n') if line.strip()]
        valid_urls = [url for url in urls if self._is_valid_url(url)]

        if not valid_urls:
            return

        self._error_placeholder.visible = False
        self._message_placeholder.param.update(
            object=f"Preparing to download {len(valid_urls)} file(s)...",
            visible=True
        )

        if self._active_download_task and not self._active_download_task.done():
            self._active_download_task.cancel()

        self._active_download_task = asyncio.create_task(self._download_and_process_urls(valid_urls))
        self._active_download_task.add_done_callback(
            lambda t: self._message_placeholder.param.update(
                object="Download complete." if not t.cancelled() else "Download cancelled.",
                visible=True
            ) if not t.cancelled() or t.exception() is None else None
        )

    async def _download_and_process_urls(self, urls):
        """Download URLs and generate file cards for processing"""
        downloaded_files = {}
        errors = []

        try:
            for i, url in enumerate(urls, 1):
                if asyncio.current_task().cancelled():
                    break

                self._message_placeholder.object = f"Processing {i} of {len(urls)}: {url[:50]}{'...' if len(url) > 50 else ''}"
                filename, content, error = await self._download_file(url)

                if error:
                    if "cancelled" in error.lower():
                        break
                    errors.append(f"{url}: {error}")
                else:
                    downloaded_files[filename] = content

        except asyncio.CancelledError:
            self._progress_bar.visible = False
            self._progress_description.visible = False
            raise

        # Generate file cards from downloaded files
        self._generate_file_cards(downloaded_files)

        if errors:
            error_msg = "\n".join(errors)
            self._error_placeholder.param.update(
                object=f"Download errors:\n{error_msg}",
                visible=True
            )

        success_count = len(downloaded_files)
        if success_count > 0:
            self._message_placeholder.object = f"Successfully downloaded {success_count} file(s). Click 'Confirm file(s)' to process."
        else:
            self._message_placeholder.visible = False

        # Clear the active task so Cancel button hides
        self._active_download_task = None

    @param.depends("add", watch=True)
    def _on_add(self):
        """Process downloaded files."""
        if len(self._file_cards) == 0:
            return

        # Clear the active task so Cancel button hides
        self._active_download_task = None

        with self._layout.param.update(loading=True):
            n_tables, n_docs, n_metadata = self._process_files()

            total_files = len(self._file_cards)
            if self.clear_uploads:
                self._clear_uploads()
                self._url_input.value = ""

            if (n_tables + n_docs + n_metadata) > 0:
                self._message_placeholder.param.update(
                    object=f"Successfully processed {total_files} files ({n_tables} table(s), {n_metadata} metadata file(s)).",
                    visible=True,
                )
            self._error_placeholder.object = self._error_placeholder.object.strip()

        self._count += 1

        if (n_tables + n_docs + n_metadata) > 0:
            self.param.trigger('upload_successful')

    def __panel__(self):
        return self._layout


class RevisionControls(Viewer):

    active = param.Boolean(False, doc="Click to revise")

    instruction = param.String(doc="Instruction to LLM to revise output")

    interface = param.Parameter()

    layout_kwargs = param.Dict(default={})

    task = param.Parameter()

    view = param.Parameter(doc="The View to revise")

    toggle_kwargs = {}

    input_kwargs = {}

    def __init__(self, **params):
        super().__init__(**params)
        input_kwargs = dict(
            max_length=200,
            margin=10,
            size="small"
        )
        input_kwargs.update(self.input_kwargs)
        self._text_input = TextInput(**input_kwargs)
        popup = Popup(
            self._text_input,
            open=self.param.active,
            anchor_origin={"horizontal": "right", "vertical": "center"},
            styles={"z-index": '1300'}
        )
        icon = ToggleIcon.from_param(
            self.param.active,
            active_icon="cancel",
            attached=[popup],
            label="",
            margin=(5, 0),
            size="small",
            icon_size="1.1em",
            **self.toggle_kwargs
        )
        popup.param.watch(self._close, "open")
        self._text_input.param.watch(self._enter_reason, "enter_pressed")
        self._row = Row(icon, **self.layout_kwargs)

    @param.depends("active", watch=True)
    def _open(self):
        if self.active:
            state.execute(self._text_input.focus, schedule=True)

    def _close(self, event):
        self.active = event.new

    def _enter_reason(self, _):
        instruction = self._text_input.value_input
        self._text_input.value = ""
        self.param.update(active=False, instruction=instruction)

    def _report_status(self, out: str, title='➕/➖ Applied Revisions'):  # noqa: RUF001
        if not out:
            return
        md = Markdown(
            out,
            margin=0,
            styles={"padding-inline": "0", "margin-left": "0", "padding": "0"},
            stylesheets=[".codehilite { margin-top: 0; margin-bottom: 0; display: inline-block; } .codehilite pre { margin-top: -1.5em; }"],
            sizing_mode="stretch_width"
        )
        self.interface.stream(Card(md, title=title), user="Assistant")

    def __panel__(self):
        return self._row


class RetryControls(RevisionControls):

    instruction = param.String(doc="Reason for retry")

    input_kwargs = {"placeholder": "Enter feedback and press the <Enter> to retry."}

    toggle_kwargs = {"icon": "auto_awesome", "description": "Prompt LLM to retry"}

    @param.depends("instruction", watch=True)
    async def _revise(self):
        if not self.instruction:
            return
        elif not hasattr(self.task.actor, 'revise'):
            invalidation_keys = set(self.task.actor.output_schema.__annotations__)
            self.task.invalidate(invalidation_keys, start=1)
            root = self.task
            while root.parent is not None:
                root = root.parent
            with root.param.update(interface=self.interface):
                await root.execute()
            return

        old_spec = self.view.spec
        messages = list(self.task.history)
        try:
            with self.task.actor.param.update(interface=self.interface), self.view.editor.param.update(loading=True):
                new_spec = await self.task.actor.revise(
                    self.instruction, messages, self.task.out_context, self.view
                )
        except Exception as e:
            self._report_status(f"```\n{e}\n```", title="❌ Failed to generate revisions")
            raise

        diff = generate_diff(old_spec, new_spec, filename=self.view.language or "spec")
        diff_md = f'```diff\n{diff}\n```'
        try:
            self.view.spec = new_spec
        except Exception:
            self.view.spec = old_spec
            self._report_status(diff_md, title="❌ Failed to apply revisions")
            raise
        self._report_status(diff_md)


class AnnotationControls(RevisionControls):
    """Controls for adding annotations to visualizations."""

    input_kwargs = {
        "placeholder": "Describe what to annotate (e.g., 'highlight peak values', 'mark outliers')..."
    }

    toggle_kwargs = {
        "description": "Add annotations to highlight key insights",
        "icon": "chat-bubble",
    }

    @param.depends("instruction", watch=True)
    async def _annotate(self):
        if not self.instruction:
            return

        old_spec = self.view.spec
        try:
            with self.task.actor.param.update(interface=self.interface), self.view.editor.param.update(loading=True):
                new_spec = await self.task.actor.annotate(
                    self.instruction, list(self.task.history), self.task.out_context, {"spec": load_yaml(self.view.spec)}
                )
        except Exception as e:
            self._report_status(f"```\n{e}\n```", title="❌ Failed to generate annotations")
            raise

        diff = generate_diff(old_spec, new_spec, filename=self.view.language or "spec")
        diff_md = f'```diff\n{diff}\n```'
        try:
            self.view.spec = new_spec
        except Exception:
            self.view.spec = old_spec
            self._report_status(diff_md, title="❌ Failed to apply annotations")
            raise
        self._report_status(diff_md)


class CopyControls(Viewer):
    """Controls for copying spec to clipboard."""

    interface = param.Parameter()

    layout_kwargs = param.Dict(default={})

    task = param.Parameter()

    view = param.Parameter(doc="The View to copy")

    def __init__(self, **params):
        super().__init__(**params)
        copy_icon = IconButton(
            icon="content_copy",
            active_icon="check",
            margin=(5, 0),
            toggle_duration=1000,
            description="Copy YAML to clipboard",
            size="small",
            color="primary",
            icon_size="0.9em"
        )
        copy_icon.js_on_click(
            args={"code_editor": self.view.editor},
            code="navigator.clipboard.writeText(code_editor.code);",
        )
        self._row = Row(copy_icon, **self.layout_kwargs)

    def __panel__(self):
        return self._row


class SourceCatalog(Viewer):
    """
    A component that displays all data sources and documents as hierarchical trees.

    Structure:
        == Sources Tree ==
        Source (with delete action)
        └── Table (checkbox for visibility)
             └── metadata_file.md (checkbox for table association)

        == Apply to All Tables ==
        ☑ readme.md (UI shortcut to check/uncheck doc under ALL tables)
        ☐ schema.md

    Document visibility:
    - Checked under table: Included when that table is relevant to query
    - Checked in "Apply to All Tables": Associates with ALL tables at once
    - Unchecked everywhere: Not included in LLM context
    """

    context = param.Dict(default={})

    sources = param.List(default=[], doc="""
        List of data sources to display in the catalog.""")

    vector_store = param.Parameter(default=None, doc="""
        The vector store to sync metadata documents to. If not provided,
        metadata files will only be stored locally.""")

    visibility_changed = param.Event(doc="""
        Triggered when table or document visibility changes.""")

    def __init__(self, /, context: TContext | None = None, **params):
        if context is None:
            raise ValueError("SourceCatalog must be given a context dictionary.")
        if "source" in context and "sources" not in context:
            context["sources"] = [context["source"]]

        super().__init__(context=context, **params)

        # === Sources Tree ===
        self._sources_title = Markdown("**Data Sources**", margin=(0, 10))
        self._sources_tree = Tree(
            items=[],
            checkboxes=True,
            propagate_to_child=True,
            color="primary",
            sizing_mode="stretch_width",
            margin=(0, 10),
        )
        self._sources_tree.param.watch(self._on_sources_active_change, "active")
        self._sources_tree.on_action("Delete", self._on_delete_source)

        # === Global Documents Tree ===
        self._docs_title = Markdown("**Apply to All Tables**", margin=(10, 10, 0, 10), visible=False)
        self._docs_tree = Tree(
            items=[],
            checkboxes=True,
            color="secondary",
            sizing_mode="stretch_width",
            margin=(0, 10),
            visible=False,
        )
        self._docs_tree.param.watch(self._on_docs_active_change, "active")

        # Combined layout
        self._layout = Column(
            self._docs_title,
            self._docs_tree,
            self._sources_title,
            self._sources_tree,
            sizing_mode="stretch_width",
            margin=(0, 0, 10, 0)
        )

        # Track the mapping from tree paths to source/table/metadata
        self._path_map = {}  # path tuple -> {"source": ..., "table": ..., "metadata": ...}
        self._docs_path_map = {}  # index -> filename
        self._suppress_sources_callback = False
        self._suppress_docs_callback = False

        # Store available metadata files
        self._available_metadata = []

    def _build_sources_items(self, sources: list) -> list[dict]:
        """
        Build the tree items structure from sources.

        Returns:
            List of item dicts for the Sources Tree component.
        """
        items = []
        available_metadata = self._available_metadata
        multiple_sources = len(sources) > 1
        self._path_map.clear()

        for src_idx, source in enumerate(sources):
            tables = source.get_tables()

            # Build table items
            table_items = []
            for tbl_idx, table in enumerate(tables):
                # Get table metadata for secondary text
                table_meta = source.metadata.get(table, {}) if source.metadata else {}
                secondary_parts = [
                    f"{k}: {v}" for k, v in table_meta.items()
                    if k not in ("docs", "columns")
                ]

                # Build metadata file items as children of this table
                metadata_items = []
                for meta_idx, meta in enumerate(available_metadata):
                    meta_path = (src_idx, tbl_idx, meta_idx)
                    self._path_map[meta_path] = {
                        "source": source,
                        "table": table,
                        "metadata": meta["filename"],
                    }
                    metadata_items.append({
                        "label": meta["filename"],
                        "icon": "description",
                    })

                table_path = (src_idx, tbl_idx)
                self._path_map[table_path] = {
                    "source": source,
                    "table": table,
                    "metadata": None,
                }

                table_item = {
                    "label": table,
                    "icon": "list",
                }
                if secondary_parts:
                    table_item["secondary"] = "; ".join(secondary_parts)
                if metadata_items:
                    table_item["items"] = metadata_items
                table_items.append(table_item)

            # Source-level path
            source_path = (src_idx,)
            self._path_map[source_path] = {
                "source": source,
                "table": None,
                "metadata": None,
            }

            source_item = {
                "label": source.name,
                "icon": "source",
            }
            if multiple_sources:
                source_item["actions"] = [{"label": "Delete", "icon": "delete"}]
            if table_items:
                source_item["items"] = table_items
            items.append(source_item)

        return items

    def _build_docs_items(self) -> list[dict]:
        """
        Build the tree items structure for global documents.

        Returns:
            List of item dicts for the Docs Tree component.
        """
        items = []
        self._docs_path_map.clear()

        for idx, meta in enumerate(self._available_metadata):
            self._docs_path_map[idx] = meta["filename"]
            items.append({
                "label": meta["filename"],
                "icon": "description",
            })

        return items

    def _compute_sources_active_paths(self, sources: list) -> list[tuple]:
        """
        Compute which paths should be active in sources tree.

        Returns paths for:
        - Tables: Only if in visible_slugs (user explicitly checked them)
        - Docs: If associated with their parent table
        - Sources: If ALL their tables are in visible_slugs
        """
        active = []
        visible_slugs = self.context.get("visible_slugs", [])
        available_metadata = self._available_metadata
        meta_filenames = [m["filename"] for m in available_metadata]

        for src_idx, source in enumerate(sources):
            tables = source.get_tables()
            all_tables_visible = True

            for tbl_idx, table in enumerate(tables):
                table_slug = f"{source.name}{SOURCE_TABLE_SEPARATOR}{table}"
                table_is_visible = table_slug in visible_slugs

                # Table is checked ONLY if in visible_slugs
                if table_is_visible:
                    active.append((src_idx, tbl_idx))
                else:
                    all_tables_visible = False

                # Add doc associations (independent of table visibility)
                if source.metadata:
                    associations = source.metadata.get(table, {}).get("docs", [])
                    for meta_filename in associations:
                        if meta_filename in meta_filenames:
                            meta_idx = meta_filenames.index(meta_filename)
                            active.append((src_idx, tbl_idx, meta_idx))

            # Source is checked ONLY if ALL tables are visible
            if all_tables_visible and len(tables) > 0:
                active.append((src_idx,))

        return active

    def _compute_docs_active_paths(self) -> list[tuple]:
        """
        Compute which paths should be active in docs tree.

        A doc is "globally checked" if it's associated with ALL tables.
        """
        sources = self.sources or self.context.get("sources", [])
        if not sources:
            return []

        # Count total tables
        total_tables = sum(len(source.get_tables()) for source in sources)
        if total_tables == 0:
            return []

        active = []
        for idx, meta in enumerate(self._available_metadata):
            filename = meta["filename"]

            # Count how many tables have this doc associated
            association_count = 0
            for source in sources:
                if not source.metadata:
                    continue
                for table in source.get_tables():
                    if table in source.metadata:
                        if filename in source.metadata[table].get("docs", []):
                            association_count += 1

            # Checked if associated with ALL tables
            if association_count == total_tables:
                active.append((idx,))

        return active

    def _compute_expanded_paths(self, sources: list) -> list[tuple]:
        """
        Compute which paths should be expanded (all sources expanded by default).
        """
        return [(src_idx,) for src_idx in range(len(sources))]

    def _on_sources_active_change(self, event):
        """Handle sources tree selection changes."""
        if self._suppress_sources_callback:
            return

        active_paths = set(event.new)

        for path, info in self._path_map.items():
            if info["table"] is None:
                continue  # Skip source-level paths

            if info["metadata"] is None:
                self._update_table_visibility(path, info, active_paths)
            else:
                self._update_table_doc_association(path, info, active_paths)

        # Re-sync the docs tree to reflect changes in doc associations
        self._sync_docs_tree()

        # Trigger visibility changed event to rebuild metaset
        self.param.trigger('visibility_changed')

    def _update_table_visibility(self, path: tuple, info: dict, active_paths: set):
        """Update table visibility based on checkbox state."""
        source = info["source"]
        table = info["table"]
        table_slug = f"{source.name}{SOURCE_TABLE_SEPARATOR}{table}"

        if path in active_paths:
            self.context["visible_slugs"].add(table_slug)
        else:
            self.context["visible_slugs"].discard(table_slug)

    def _update_table_doc_association(self, path: tuple, info: dict, active_paths: set):
        """Update document-table association based on checkbox state."""
        source = info["source"]
        table = info["table"]
        metadata = info["metadata"]

        self._ensure_source_metadata(source)
        associations = self._get_table_docs(source, table)

        if path in active_paths:
            if metadata not in associations:
                associations.append(metadata)
        elif metadata in associations:
            associations.remove(metadata)

    def _on_docs_active_change(self, event):
        """
        Handle docs tree selection changes.

        Global toggle is just a UI convenience:
        - Check → Associate doc with ALL tables
        - Uncheck → Remove doc from ALL tables
        """
        if self._suppress_docs_callback:
            return

        active_indices = {path[0] for path in event.new if len(path) == 1}
        sources = self.sources or self.context.get("sources", [])

        # Track visible docs in context (mirrors visible_slugs pattern)
        if "visible_docs" not in self.context:
            self.context["visible_docs"] = set()

        for idx, filename in self._docs_path_map.items():
            is_active = idx in active_indices
            if is_active:
                self._associate_doc_with_all_tables(filename, sources)
                self.context["visible_docs"].add(filename)
            else:
                self._remove_doc_from_all_tables(filename, sources)
                self.context["visible_docs"].discard(filename)

        # Re-sync the sources tree to reflect the updated associations
        self._sync_sources_tree_only()

        # Trigger visibility changed event to rebuild metaset
        self.param.trigger('visibility_changed')

    def _associate_doc_with_all_tables(self, filename: str, sources: list):
        """Associate doc with all tables."""
        for source in sources:
            self._ensure_source_metadata(source)
            for table in source.get_tables():
                associations = self._get_table_docs(source, table)
                if filename not in associations:
                    associations.append(filename)

    def _remove_doc_from_all_tables(self, filename: str, sources: list):
        """Remove doc from all tables."""
        for source in sources:
            if not source.metadata:
                continue

            for table in source.get_tables():
                if table not in source.metadata or "docs" not in source.metadata[table]:
                    continue

                associations = source.metadata[table]["docs"]
                if filename in associations:
                    associations.remove(filename)

    def _ensure_source_metadata(self, source):
        """Ensure source has metadata structure initialized."""
        if source.metadata is None:
            source.metadata = {}

    def _get_table_docs(self, source, table: str) -> list:
        """Get or create the docs list for a table."""
        if table not in source.metadata:
            source.metadata[table] = {}
        if "docs" not in source.metadata[table]:
            source.metadata[table]["docs"] = []
        return source.metadata[table]["docs"]

    def _on_delete_source(self, item: dict):
        """Handle delete action on a source node."""
        source_name = item["label"]
        sources = self.context.get("sources", [])

        source_to_delete = self._find_source_by_name(sources, source_name)
        if not source_to_delete:
            return

        self._remove_source_tables_from_visible_slugs(source_to_delete)
        self._remove_source_from_context(sources, source_to_delete)

    def _find_source_by_name(self, sources: list, source_name: str):
        """Find source by name, return None if not found."""
        for source in sources:
            if source.name == source_name:
                return source
        return None

    def _remove_source_tables_from_visible_slugs(self, source):
        """Remove all tables from a source from visible_slugs."""
        for table in source.get_tables():
            table_slug = f"{source.name}{SOURCE_TABLE_SEPARATOR}{table}"
            self.context["visible_slugs"].discard(table_slug)

    def _remove_source_from_context(self, sources: list, source_to_delete):
        """Remove source from context and trigger sync."""
        self.context["sources"] = [s for s in sources if s is not source_to_delete]
        self.sources = self.context["sources"]

    async def _sync_metadata_to_vector_store(self, filename: str):
        """
        Add metadata file to the vector store.

        This ensures document chunks are queryable alongside tables.
        """
        if self.vector_store is None:
            return

        available = {
            m["filename"]: m for m in self._available_metadata
        }

        if filename not in available:
            return

        metadata_info = available[filename]
        content = metadata_info.get("content", "")

        doc_entry = {
            "text": content,
            "metadata": {
                "filename": filename,
                "type": "document",
            },
        }

        try:
            await self.vector_store.upsert([doc_entry])
        except Exception:
            raise

    @param.depends("sources", watch=True)
    def _on_sources_change(self):
        """Trigger sync when sources param changes."""
        self.sync()

    def _sync_sources_tree_only(self):
        """Sync only the sources tree, without touching the docs tree."""
        sources = self.sources or self.context.get("sources", [])

        if not sources:
            return

        items = self._build_sources_items(sources)
        if not items:
            return

        self._update_sources_tree_state(items, sources)

    def sync(self, context: TContext | None = None):
        """
        Synchronize both trees with current sources and metadata.

        Args:
            context: Optional context dict. If None, uses self.context.
        """
        context = context or self.context
        sources = self.sources or context.get("sources", [])

        self._sync_sources_tree(sources)
        self._sync_docs_tree()

        self._layout.loading = False

    def _sync_sources_tree(self, sources: list):
        """Sync the sources tree with current sources."""
        if not sources:
            self._sources_title.object = "**Available Sources** *(no sources available)*"
            self._sources_tree.items = []
            return

        items = self._build_sources_items(sources)
        if not items:
            self._sources_title.object = "**Available Sources** *(no tables found)*"
            self._sources_tree.items = []
            return

        self._sources_title.object = "**Available Sources**"
        self._auto_associate_unassociated_metadata(sources)
        self._update_sources_tree_state(items, sources)

    def _auto_associate_unassociated_metadata(self, sources: list):
        """Automatically associate metadata files that aren't yet associated with any table.
        Also marks tables as visible on initial upload and initializes visible_docs."""
        meta_filenames = [m["filename"] for m in self._available_metadata]
        associated = self._collect_associated_metadata(sources)
        unassociated = set(meta_filenames) - associated

        if not unassociated:
            return

        # Initialize visible_docs if not present
        if "visible_docs" not in self.context:
            self.context["visible_docs"] = set()

        # Associate with all tables and mark them as visible
        for source in sources:
            for table in source.get_tables():
                table_slug = f"{source.name}{SOURCE_TABLE_SEPARATOR}{table}"

                self._ensure_source_metadata(source)
                associations = self._get_table_docs(source, table)

                for meta_filename in unassociated:
                    if meta_filename not in associations:
                        associations.append(meta_filename)

                # Mark table as visible (only happens on initial upload)
                self.context["visible_slugs"].add(table_slug)

        # Mark unassociated docs as visible (they're being auto-associated)
        for meta_filename in unassociated:
            self.context["visible_docs"].add(meta_filename)

        # Sync to vector store once per document (not once per table)
        for meta_filename in unassociated:
            asyncio.create_task(self._sync_metadata_to_vector_store(meta_filename))  # noqa: RUF006

    def _collect_associated_metadata(self, sources: list) -> set:
        """Collect all metadata filenames currently associated with any table."""
        associated = set()
        for source in sources:
            if not source.metadata:
                continue
            for table in source.get_tables():
                if table in source.metadata:
                    associations = source.metadata[table].get("docs", [])
                    associated.update(associations)
        return associated

    def _update_sources_tree_state(self, items: list, sources: list):
        """Update sources tree items, active paths, and expanded paths."""
        active = self._compute_sources_active_paths(sources)
        expanded = self._compute_expanded_paths(sources)

        self._suppress_sources_callback = True
        try:
            self._sources_tree.items = items
            self._sources_tree.active = active
            self._sources_tree.expanded = expanded
        finally:
            self._suppress_sources_callback = False

    def _sync_docs_tree(self):
        """Sync the global docs tree with available metadata."""
        if not self._available_metadata:
            self._docs_title.visible = False
            self._docs_tree.visible = False
            return

        self._docs_title.visible = True
        self._docs_title.object = "**Apply to All Tables**"
        self._docs_tree.visible = True

        docs_items = self._build_docs_items()
        docs_active = self._compute_docs_active_paths()

        self._suppress_docs_callback = True
        try:
            self._docs_tree.items = docs_items
            self._docs_tree.active = docs_active
        finally:
            self._suppress_docs_callback = False

    def __panel__(self):
        """
        Return the tree layout.
        """
        return self._layout



class TableExplorer(Viewer):
    """
    TableExplorer provides a high-level entrypoint to explore tables in a split UI.
    It allows users to load tables, explore them using Graphic Walker, and then
    interrogate the data via a chat interface.
    """

    add_exploration = param.Event(label="Explore table")

    table_slug = param.Selector(label="Select table(s) to preview.")

    context = param.Dict(default={})

    def __init__(self, **params):
        self._initialized = False
        super().__init__(**params)
        self._table_autocomplete = AutocompleteInput(
            value="",
            options=[],
            restrict=True,
            case_sensitive=False,
            search_strategy="includes",
            min_characters=0,
            placeholder="Select or search for a table...",
            sizing_mode='stretch_width',
        )
        self._table_autocomplete.param.watch(self._on_table_selected, 'value')
        self._input_row = Row(
            self._table_autocomplete
        )
        self.source_map = {}
        self._layout = self._input_row

    def _on_table_selected(self, event):
        """Handle table selection from autocomplete."""
        if event.new:
            self.table_slug = event.new
            self.param.trigger('add_exploration')
            # Clear the autocomplete after selection
            self._table_autocomplete.value = ""

    @param.depends("context", watch=True)
    async def sync(self, context: TContext | None = None):
        init = not self._initialized
        self._initialized = True
        context = context or self.context
        if "sources" in context:
            sources = context["sources"]
        elif "source" in context:
            sources = [context["source"]]
        else:
            return
        selected = [self.table_slug] if self.table_slug else []
        deduplicate = len(sources) > 1
        new = {}

        # Build the source map for UI display
        for source in sources:
            tables = source.get_tables()
            for table in tables:
                if deduplicate:
                    table = f'{source.name}{SOURCE_TABLE_SEPARATOR}{table}'

                if (table.split(SOURCE_TABLE_SEPARATOR, maxsplit=1)[-1] not in self.source_map and
                    not init and state.loaded):
                    selected.append(table)
                new[table] = source

        self.source_map.clear()
        self.source_map.update(new)
        selected = selected[-1] if len(selected) == 1 else None

        self._table_autocomplete.options = list(self.source_map)
        if selected:
            self.table_slug = selected
        self._input_row.visible = bool(self.source_map)
        self._initialized = True

    def create_sql_output(self) -> SQLOutput | None:
        if not self.table_slug:
            return
        from .views import SQLOutput

        source = self.source_map[self.table_slug]
        if SOURCE_TABLE_SEPARATOR in self.table_slug:
            _, table = self.table_slug.split(SOURCE_TABLE_SEPARATOR, maxsplit=1)
        else:
            table = self.table_slug
        new_table = f"select_{table}"
        sql_expr = f"SELECT * FROM \"{table}\""
        new_source = source.create_sql_expr_source({new_table: sql_expr})
        pipeline = Pipeline(source=new_source, table=new_table)
        return SQLOutput(spec=sql_expr, component=pipeline)

    def __panel__(self):
        return self._layout
